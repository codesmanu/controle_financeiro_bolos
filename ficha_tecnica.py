"""
Lógica compartilhada para montar a aba "Ficha Técnica e Custos".

Por que este arquivo existe
---------------------------
gerar_planilha.py e gerar_dashboard.py montavam essa aba cada um do seu
jeito, com o código quase idêntico copiado e colado — inclusive com o
nome da aba escrito diferente em cada um ("1. Ficha Técnica" vs
"1. Ficha Técnica e Custos"). Isso fazia o atualizar_planilha.py só
conseguir enxergar um dos dois arquivos gerados.

Agora só existe uma função (montar_ficha_tecnica) que monta a aba, usada
pelos dois scripts. Assim os dois ficam sempre iguais e com o mesmo nome
de aba.
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from precos import carregar_precos

NOME_ABA_FICHA_TECNICA = "1. Ficha Técnica e Custos"

HEADER_FILL = PatternFill(start_color="4A2E2B", end_color="4A2E2B", fill_type="solid")
SUB_FILL = PatternFill(start_color="D9C3B0", end_color="D9C3B0", fill_type="solid")
ZEBRA_FILL = PatternFill(start_color="F9F6F0", end_color="F9F6F0", fill_type="solid")
GREEN_FILL = PatternFill(start_color="E8F0E6", end_color="E8F0E6", fill_type="solid")

FONT_TITLE = Font(name="Arial", size=15, bold=True, color="4A2E2B")
FONT_HEADER = Font(name="Arial", size=11, bold=True, color="FFFFFF")
FONT_BOLD = Font(name="Arial", size=11, bold=True, color="000000")
FONT_PROFIT = Font(name="Arial", size=11, bold=True, color="2E5A27")

_BORDER_THIN = Side(border_style="thin", color="CCCCCC")
BOX_BORDER = Border(left=_BORDER_THIN, right=_BORDER_THIN, top=_BORDER_THIN, bottom=_BORDER_THIN)


def _item(ingredientes, chave, uso, formula):
    """Monta a linha (nome, embalagem, preço, uso na receita, fórmula)
    de um ingrediente a partir do precos.json."""
    ing = ingredientes[chave]
    return (ing["nome"], ing["embalagem"], ing["preco"], uso, formula)


def montar_ficha_tecnica(ws):
    """Preenche a planilha `ws` (já criada/renomeada pelo script chamador)
    com a ficha técnica completa: ingredientes, preços, fórmulas de custo
    e o resumo final por potinho."""
    ingredientes = carregar_precos()
    ws.views.sheetView[0].showGridLines = True

    ws["A1"] = "🍰 FICHA TÉCNICA E CUSTO AUTOMÁTICO (250g)"
    ws["A1"].font = FONT_TITLE

    headers = ["Ingrediente / Etapa", "Embalagem", "Preço Pago (R$)", "Uso na Receita", "Custo (R$)"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.font, c.fill, c.alignment = FONT_HEADER, HEADER_FILL, Alignment(horizontal="center")

    dados = [
        _item(ingredientes, "farinha_trigo", "360g", "=C5*(360/1000)"),
        _item(ingredientes, "acucar", "240g", "=C6*(240/1000)"),
        _item(ingredientes, "margarina", "100g", "=C7*(100/250)"),
        _item(ingredientes, "leite_liquido", "300ml", "=C8*(300/1000)"),
        _item(ingredientes, "fermento", "10g", "=C9*(10/100)"),
        _item(ingredientes, "ovos", "3 un", "=C10*3"),
        ("👉 CUSTO DA MASSA (POR BOLO)", "", "", "1 un (1/21 da massa)", "=SUM(E5:E10)/21"),
        _item(ingredientes, "leite_condensado", "3 caixas", "=C12"),
        _item(ingredientes, "creme_de_leite", "3 caixas", "=C13"),
        _item(ingredientes, "chocolate_po", "45g", "=C14*(45/1000)"),
        _item(ingredientes, "coco_ralado", "50g", "=C15"),
        _item(ingredientes, "leite_em_po", "30g", "=C16*(30/200)"),
        ("👉 CUSTO DO RECHEIO (POR BOLO)", "", "", "1 un (1/11 dos recheios)", "=SUM(E12:E16)/11"),
        _item(ingredientes, "calda_refrigerante", "Porção (~1/50)", "=C18/50"),
        _item(ingredientes, "kit_embalagem", "1 kit", "=C19"),
    ]

    for r_idx, linha in enumerate(dados, start=5):
        is_destaque = "👉" in linha[0]
        for c_idx, val in enumerate(linha, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = BOX_BORDER
            if is_destaque:
                cell.font, cell.fill = FONT_BOLD, SUB_FILL
            elif r_idx % 2 == 0:
                cell.fill = ZEBRA_FILL
            if c_idx in [3, 5] and val != "":
                cell.number_format = "R$ #,##0.00"
            if c_idx in [2, 4]:
                cell.alignment = Alignment(horizontal="center")

    ws["G4"] = "💡 RESUMO FINAL POR POTINHO"
    ws["G4"].font = Font(name="Arial", size=11, bold=True, color="4A2E2B")

    resumo = [
        ("Soma Custo Direto:", "=E11+E17+E18+E19"),
        ("+ Gás/Luz/Invisíveis (10%):", "=I5*0.1"),
        ("🔥 CUSTO TOTAL:", "=I5+I6"),
        ("Lucro Venda R$ 12,00:", "=12-I7"),
        ("Lucro Venda R$ 13,00:", "=13-I7"),
    ]
    for idx, (label, form) in enumerate(resumo, start=5):
        ws.cell(row=idx, column=7, value=label).font = FONT_BOLD if ("TOTAL" in label or "Lucro" in label) else Font(name="Arial", size=11)
        c_val = ws.cell(row=idx, column=9, value=form)
        c_val.number_format = "R$ #,##0.00"
        c_val.font = FONT_PROFIT if "Lucro" in label else FONT_BOLD
        if "Lucro" in label:
            ws.cell(row=idx, column=7).fill, c_val.fill = GREEN_FILL, GREEN_FILL
        elif "TOTAL" in label:
            ws.cell(row=idx, column=7).fill, c_val.fill = SUB_FILL, SUB_FILL