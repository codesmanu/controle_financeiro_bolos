import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

from ficha_tecnica import (
    montar_ficha_tecnica,
    NOME_ABA_FICHA_TECNICA,
    FONT_TITLE,
    FONT_HEADER,
    FONT_BOLD,
    FONT_PROFIT,
    HEADER_FILL,
    GREEN_FILL,
    BOX_BORDER,
)

wb = openpyxl.Workbook()


# ABA 1: FICHA TÉCNICA (mesma função usada pelo gerar_planilha.py, garantindo
# que as duas planilhas fiquem sempre iguais e com o mesmo nome de aba)

ws1 = wb.active
ws1.title = NOME_ABA_FICHA_TECNICA
montar_ficha_tecnica(ws1)


# ABA 2: CONTROLE DE VENDAS

ws2 = wb.create_sheet(title="2. Controle de Vendas")
ws2.views.sheetView[0].showGridLines = True
ws2["A1"] = "📊 DIÁRIO DE VENDAS E REGISTRO DE PIX"
ws2["A1"].font = FONT_TITLE

headers2 = ["Data", "Cliente", "Local", "Qtd", "Preço (R$)", "Total (R$)", "Custo (R$)", "LUCRO LIMPO (R$)"]
for col_idx, h in enumerate(headers2, start=1):
    c = ws2.cell(row=4, column=col_idx, value=h)
    c.font, c.fill, c.alignment = FONT_HEADER, HEADER_FILL, Alignment(horizontal="center")

vendas_ex = [
    ("14/07/2026", "Andrea Rodrigues", "Casa", 1, 12.00),
    ("14/07/2026", "Maria Eduarda", "Casa", 1, 12.00),
    ("15/07/2026", "Marcus Wilker (Irmão)", "Trabalho", 3, 13.00),
    ("16/07/2026", "Kamilly Vitória", "Casa", 1, 12.00),
    ("17/07/2026", "Franklimilton", "Casa", 1, 12.00),
    ("19/07/2026", "Franklimilton", "Casa", 2, 12.00),
    ("20/07/2026", "Marcus Wilker (Irmão)", "Trabalho", 4, 13.00),
    ("20/07/2026", "Marcus Wilker (Irmão)", "Trabalho", 5, 13.00),
]

LINHA_INICIAL = 5
LINHA_ULTIMO_EXEMPLO = LINHA_INICIAL + len(vendas_ex) - 1

# Linhas em branco já formatadas e com fórmula pronta, reservadas para novas
# vendas que você for registrando com o tempo — assim a linha de TOTAL não
# fica "presa" logo depois dos 8 exemplos.
N_LINHAS_RESERVADAS = 40
LINHA_ULTIMA_VENDA = LINHA_ULTIMO_EXEMPLO + N_LINHAS_RESERVADAS
LINHA_TOTAIS = LINHA_ULTIMA_VENDA + 1

for idx, (dt, cli, loc, qtd, prc) in enumerate(vendas_ex, start=LINHA_INICIAL):
    ws2.cell(row=idx, column=1, value=dt).alignment = Alignment(horizontal="center")
    ws2.cell(row=idx, column=2, value=cli)
    ws2.cell(row=idx, column=3, value=loc).alignment = Alignment(horizontal="center")
    ws2.cell(row=idx, column=4, value=qtd).alignment = Alignment(horizontal="center")
    ws2.cell(row=idx, column=5, value=prc).number_format = "R$ #,##0.00"

    ws2.cell(row=idx, column=6, value=f"=D{idx}*E{idx}").number_format = "R$ #,##0.00"
    ws2.cell(row=idx, column=7, value=f"=D{idx}*'{NOME_ABA_FICHA_TECNICA}'!$I$7").number_format = "R$ #,##0.00"

    c_prof = ws2.cell(row=idx, column=8, value=f"=F{idx}-G{idx}")
    c_prof.number_format = "R$ #,##0.00"
    c_prof.font, c_prof.fill = FONT_PROFIT, GREEN_FILL

    for c in range(1, 9):
        ws2.cell(row=idx, column=c).border = BOX_BORDER

# Linhas reservadas para novas vendas: ficam em branco (colunas A-E), mas já
# calculam Total/Custo/Lucro sozinhas assim que você preencher Qtd e Preço.
for idx in range(LINHA_ULTIMO_EXEMPLO + 1, LINHA_ULTIMA_VENDA + 1):
    ws2.cell(row=idx, column=6, value=f'=IF(D{idx}="","",D{idx}*E{idx})').number_format = "R$ #,##0.00"
    ws2.cell(row=idx, column=7, value=f'=IF(D{idx}="","",D{idx}*\'{NOME_ABA_FICHA_TECNICA}\'!$I$7)').number_format = "R$ #,##0.00"

    c_prof = ws2.cell(row=idx, column=8, value=f'=IF(D{idx}="","",F{idx}-G{idx})')
    c_prof.number_format = "R$ #,##0.00"
    c_prof.font, c_prof.fill = FONT_PROFIT, GREEN_FILL

    for c in range(1, 9):
        ws2.cell(row=idx, column=c).border = BOX_BORDER

# Linha de Totais do Mês
ws2.cell(row=LINHA_TOTAIS, column=3, value="TOTAL DO MÊS:").font = FONT_BOLD
ws2.cell(row=LINHA_TOTAIS, column=4, value=f"=SUM(D{LINHA_INICIAL}:D{LINHA_ULTIMA_VENDA})").font = FONT_BOLD
ws2.cell(row=LINHA_TOTAIS, column=6, value=f"=SUM(F{LINHA_INICIAL}:F{LINHA_ULTIMA_VENDA})").font = FONT_BOLD
ws2.cell(row=LINHA_TOTAIS, column=6).number_format = "R$ #,##0.00"
ws2.cell(row=LINHA_TOTAIS, column=7, value=f"=SUM(G{LINHA_INICIAL}:G{LINHA_ULTIMA_VENDA})").font = FONT_BOLD
ws2.cell(row=LINHA_TOTAIS, column=7).number_format = "R$ #,##0.00"
ws2.cell(row=LINHA_TOTAIS, column=8, value=f"=SUM(H{LINHA_INICIAL}:H{LINHA_ULTIMA_VENDA})").font = FONT_PROFIT
ws2.cell(row=LINHA_TOTAIS, column=8).number_format = "R$ #,##0.00"
ws2.cell(row=LINHA_TOTAIS, column=8).fill = GREEN_FILL


# ABA 3: DASHBOARD VISUAL COM GRÁFICO

ws3 = wb.create_sheet(title="3. Dashboard Visual")
ws3.views.sheetView[0].showGridLines = True
ws3["A1"] = "📈 DASHBOARD DE DESEMPENHO E GRÁFICOS"
ws3["A1"].font = FONT_TITLE

headers3 = ["Indicador", "Valor (R$)", "Descrição"]
for col_idx, h in enumerate(headers3, start=1):
    c = ws3.cell(row=4, column=col_idx, value=h)
    c.font, c.fill, c.alignment = FONT_HEADER, HEADER_FILL, Alignment(horizontal="center")

dash_items = [
    ("Faturamento Total de Vendas", f"='2. Controle de Vendas'!F{LINHA_TOTAIS}", "Soma de todas as vendas do mês"),
    ("Custo Total de Produção Usado", f"='2. Controle de Vendas'!G{LINHA_TOTAIS}", "Custo das matérias-primas e potes consumidos"),
    ("LUCRO LIMPO REAL", f"='2. Controle de Vendas'!H{LINHA_TOTAIS}", "Dinheiro limpo no seu bolso!"),
]

for idx, (ind, form, desc) in enumerate(dash_items, start=5):
    ws3.cell(row=idx, column=1, value=ind).font = FONT_BOLD
    c_v = ws3.cell(row=idx, column=2, value=form)
    c_v.number_format = "R$ #,##0.00"
    c_v.font = FONT_PROFIT if "LUCRO" in ind else FONT_BOLD
    ws3.cell(row=idx, column=3, value=desc)
    for c in range(1, 4):
        ws3.cell(row=idx, column=c).border = BOX_BORDER
        if "LUCRO" in ind:
            ws3.cell(row=idx, column=c).fill = GREEN_FILL


chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "Faturamento vs Custo vs Lucro Real (Julho)"
chart.y_axis.title = "Valor (R$)"
chart.x_axis.title = "Indicadores Financeiros"

data_ref = Reference(ws3, min_col=2, min_row=4, max_row=7)
cats_ref = Reference(ws3, min_col=1, min_row=5, max_row=7)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.width = 16
chart.height = 10

# Adiciona o gráfico exatamente a partir da célula E4 da Aba 3!
ws3.add_chart(chart, "E4")


for ws in [ws1, ws2, ws3]:
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)


wb.save("Minha_Planilha_Completa.xlsx")
print("✅ SUCESSO! Arquivo 'Minha_Planilha_Completa.xlsx' gerado com Dashboard e Gráficos de Barras!")