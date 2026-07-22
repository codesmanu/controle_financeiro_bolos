import openpyxl
from openpyxl.utils import get_column_letter

from ficha_tecnica import montar_ficha_tecnica, NOME_ABA_FICHA_TECNICA


wb = openpyxl.Workbook()
wb.active.title = NOME_ABA_FICHA_TECNICA
ws1 = wb[NOME_ABA_FICHA_TECNICA]

montar_ficha_tecnica(ws1)

for col in ws1.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)

ws1.column_dimensions["A"].width = 32
ws1.column_dimensions["G"].width = 28


wb.save("Minha_Planilha_Gerada.xlsx")
print("Planilha Minha_Planilha_Gerada.xlsx criada com sucesso!")


# comando no terminal: python gerar_planilha.py - escrever e abrir a planilha gerada no Excel