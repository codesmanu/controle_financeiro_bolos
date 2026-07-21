import openpyxl

# 1. Carregar a planilha existente que já criamos antes
nome_arquivo = "Minha_Planilha_Gerada.xlsx"
try:
    wb = openpyxl.load_workbook(nome_arquivo)
    ws = wb["1. Ficha Técnica"]
except FileNotFoundError:
    print(f"Erro: O arquivo '{nome_arquivo}' não foi encontrado! Rode o gerar_planilha.py primeiro.")
    exit()

print("="*55)
print(" 🍰 ATUALIZADOR AUTOMÁTICO DE PREÇOS NA PLANILHA ")
print("="*55)

# 2. Mapeamento das linhas exatas onde estão os preços na sua planilha
itens_mapeados = {
    "1": ("Farinha de Trigo (1kg)", 5),
    "2": ("Açúcar (1kg)", 6),
    "3": ("Margarina (250g)", 7),
    "4": ("Leite Líquido (1L)", 8),
    "5": ("Ovos (1 unidade)", 10),
    "6": ("Leite Condensado (3 caixas)", 12),
    "7": ("Creme de Leite (3 caixas)", 13),
    "8": ("Chocolate em Pó 50% (1kg)", 14),
    "9": ("Kit Embalagem + Delivery", 19)
}

print("\nEscolha qual ingrediente mudou de preço no mercado:")
for chave, (nome, _) in itens_mapeados.items():
    print(f"  [{chave}] {nome}")

opcao = input("\nDigite o NÚMERO do item e aperte ENTER: ")

if opcao in itens_mapeados:
    nome_item, linha = itens_mapeados[opcao]
    preco_atual = ws.cell(row=linha, column=3).value
    print(f"\n-> O preço atual de '{nome_item}' na planilha é: R$ {preco_atual:.2f}")
    
    try:
        # Pede o novo preço e substitui vírgula por ponto automaticamente
        entrada_preco = input("Digite o NOVO PREÇO pago no mercado (ex: 6.50): ").replace(",", ".")
        novo_preco = float(entrada_preco)
        
        # Altera o valor na célula exata da coluna 3 (Preço Pago)
        ws.cell(row=linha, column=3, value=novo_preco)
        
        # Salva a planilha atualizada
        wb.save(nome_arquivo)
        print(f"\n✅ SUCESSO! O preço de '{nome_item}' foi atualizado para R$ {novo_preco:.2f}!")
        print("💡 Quando você abrir a planilha no Excel, o Custo do Bolo e o Lucro já terão sido recalculados sozinhos!")
    except ValueError:
        print("❌ Valor inválido! Digite apenas números (ex: 6.50).")
else:
    print("Saindo sem alterar nada...")